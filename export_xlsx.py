import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

os.chdir('/Users/a1/cidi-cengyou/data')
files = [
    ('poems.json','①古龙井·风篁岭'),
    ('trail-gushan.json','②孤山白堤'),
    ('trail-lingyin.json','③灵隐天竺'),
    ('trail-qiantang.json','④钱塘江畔'),
    ('trail-xihu-scatter.json','⑤西湖名胜诗签'),
]

wb = openpyxl.Workbook()
HEAD_FILL = PatternFill("solid", fgColor="2F4858")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
A_FILL = PatternFill("solid", fgColor="E8F5E9")
B_FILL = PatternFill("solid", fgColor="FFF3E0")
EST_FILL = PatternFill("solid", fgColor="FFFDE7")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

def style_header(ws, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(1, c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = CENTER; cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

ws = wb.active
ws.title = "全部点位"
cols = ["诗径","布局","序","点位ID","点位名称","副标题","纬度lat","经度lng",
        "坐标系","坐标源","精度m","半径m","已核半径","年份","年份约","朝代",
        "作者","篇名","可信度","诗文","白话译","考据/公案","出处ref","坐标说明"]
ws.append(cols)
style_header(ws, len(cols))

row = 2; seq = 0
for f, label in files:
    d = json.load(open(f)); m = d.get('meta', {}); layout = m.get('layout','trail')
    for p in d.get('points', []):
        seq += 1
        poems = p.get('poems', [])
        for qi, q in enumerate(poems):
            credit = q.get('credit','')
            ws.append([
                label if qi==0 else "", layout if qi==0 else "", seq if qi==0 else "",
                p.get('id','') if qi==0 else "", p.get('name','') if qi==0 else "",
                p.get('subtitle','') if qi==0 else "",
                p.get('lat') if qi==0 else "", p.get('lng') if qi==0 else "",
                p.get('coord_system','') if qi==0 else "",
                p.get('coord_source','') if qi==0 else "",
                p.get('coord_accuracy_m') if qi==0 else "",
                p.get('radius') if qi==0 else "",
                ("是" if p.get('radius_verified') else "否") if qi==0 else "",
                p.get('year') if qi==0 else "",
                ("约" if p.get('year_approx') else "") if qi==0 else "",
                q.get('dynasty',''), q.get('author',''), q.get('source',''),
                credit, q.get('text',''), q.get('vernacular',''),
                q.get('annotation',''), q.get('ref',''),
                p.get('coord_note','') if qi==0 else "",
            ])
            cc = ws.cell(row, 19)
            if credit == 'A': cc.fill = A_FILL
            elif credit == 'B': cc.fill = B_FILL
            cc.alignment = CENTER
            if p.get('coord_source')=='estimate':
                ws.cell(row,10).fill = EST_FILL
            row += 1

widths = {"A":14,"B":8,"C":5,"D":7,"E":14,"F":16,"G":11,"H":11,"I":9,"J":9,
          "K":7,"L":7,"M":8,"N":7,"O":7,"P":8,"Q":12,"R":22,"S":8,
          "T":48,"U":48,"V":52,"W":30,"X":34}
for col,w in widths.items(): ws.column_dimensions[col].width = w
for r in range(2, row):
    for c in range(1, len(cols)+1):
        cell = ws.cell(r,c); cell.border = BORDER
        if c in (20,21,22,23,24,6): cell.alignment = WRAP
        elif c in (7,8): cell.alignment = Alignment(horizontal="center")

ws2 = wb.create_sheet("诗径概览")
cols2 = ["诗径","布局","点数","诗作数","A级","B级","osm点","estimate点","代表诗人","城市开篇"]
ws2.append(cols2)
style_header(ws2, len(cols2))
r2 = 2; tot=[0]*6
for f, label in files:
    d = json.load(open(f)); m=d.get('meta',{}); pts=d.get('points',[])
    npoems=sum(len(p.get('poems',[])) for p in pts)
    na=sum(1 for p in pts for q in p.get('poems',[]) if q.get('credit')=='A')
    nb=sum(1 for p in pts for q in p.get('poems',[]) if q.get('credit')=='B')
    nosm=sum(1 for p in pts if p.get('coord_source')=='osm')
    nest=sum(1 for p in pts if p.get('coord_source')=='estimate')
    authors=[]
    for p in pts:
        for q in p.get('poems',[]):
            a=q.get('author','')
            if a and a not in authors: authors.append(a)
    ws2.append([label, m.get('layout','trail'), len(pts), npoems, na, nb, nosm, nest,
                "、".join(authors[:6]), "三声部" if m.get('prologue') else ""])
    for i,v in enumerate([len(pts),npoems,na,nb,nosm,nest]): tot[i]+=v
    r2+=1
ws2.append(["合计","",tot[0],tot[1],tot[2],tot[3],tot[4],tot[5],"",""])
for c in range(1,len(cols2)+1):
    ws2.cell(r2,c).font=Font(bold=True); ws2.cell(r2,c).fill=PatternFill("solid",fgColor="ECEFF1")
w2={"A":16,"B":9,"C":7,"D":8,"E":7,"F":7,"G":8,"H":12,"I":40,"J":10}
for col,w in w2.items(): ws2.column_dimensions[col].width=w
for r in range(2,r2+1):
    for c in range(1,len(cols2)+1):
        cell=ws2.cell(r,c); cell.border=BORDER
        if c==9: cell.alignment=WRAP
        elif c!=1: cell.alignment=CENTER
note_r=r2+2
ws2.cell(note_r,1,"说明：A=确指此地此景此事、文献可考；B=确与此地相关但存公案/非专为此点而作。坐标系统一WGS-84(对齐手机GPS)。estimate点为史料估算，待实地打点校准。").font=Font(italic=True,size=9,color="666666")

out = '/Users/a1/cidi-cengyou/诗径-全部点位数据.xlsx'
wb.save(out)
print("已保存:", out)
print("Sheet1 全部点位明细行(含多诗分行):", row-2)
print("Sheet2 诗径:", r2-1, "条 + 合计")
print("总计 %d点位 / %d诗作 / A级%d B级%d / osm%d estimate%d" % (tot[0],tot[1],tot[2],tot[3],tot[4],tot[5]))
print("文件大小:", os.path.getsize(out), "bytes")
